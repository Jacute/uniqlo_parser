import json
import re

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from googletrans import Translator
from re import search
from datetime import datetime

import importlib.util

from config.colors import *
from config.config import *
from config.materials import *

from random import random
import shutil
import requests
import os
import logging
import time
import sys
import traceback
import argparse


class Parser:
    def __init__(self):
        self.result = []
        parser = argparse.ArgumentParser(description='Process some integers.')
        parser.add_argument('--headless', action='store_true', help='headless')
        args = parser.parse_args()
        if args.headless:
            self.driver = self.get_driver(True)
        else:
            self.driver = self.get_driver(False)

    def get_driver(self, headless):
        try:
            options = webdriver.ChromeOptions()
            if headless:
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')

            options.add_argument('--log-level=3')
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            options.add_argument(
                "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")

            # options.add_argument('--disable-dev-shm-usage')
            # options.add_argument('--no-sandbox')
            service = Service(os.path.abspath("chromedriver") if os.name == 'posix' else os.path.abspath("chromedriver.exe"))
            driver = webdriver.Chrome(
                service=service,
                options=options
            )
            driver.set_window_size(1920, 1080)
            driver.implicitly_wait(30)

            self.wait = WebDriverWait(driver, 30)

            return driver
        except Exception as e:
            print('Неудачная настройка браузера!')
            print(traceback.format_exc())
            print(input('Нажмите ENTER, чтобы закрыть эту программу'))
            sys.exit()

    def get_all_products(self):
        products = []
        # scroll page
        while True:
            scroll_height = 2000
            document_height_before = self.driver.execute_script("return document.documentElement.scrollHeight")
            self.driver.execute_script(f"window.scrollTo(0, {document_height_before + scroll_height});")
            time.sleep(1.5)
            document_height_after = self.driver.execute_script("return document.documentElement.scrollHeight")
            if document_height_after == document_height_before:
                break
        products.extend([i.get_attribute('href') for i in self.driver.find_elements(By.CSS_SELECTOR, '.productTile__link.productTile__imageContainerChild')])
        return list(set(products))

    def parse(self):
        c = 0
        self.driver.get(self.CATEGORIE_URL)

        products = self.get_all_products()

        for product_url in products[:PARSE_LIMIT]:
            print(f'{products.index(product_url) + 1} of {len(products[:PARSE_LIMIT])}')
            try:
                self.driver.get(product_url)
            except:
                continue

            self.driver.execute_script("window.scrollTo(0, 150)")

            deliverySectionItems = self.driver.find_elements(By.CSS_SELECTOR, '.textToggle__text.js-toggleText')

            if len(deliverySectionItems) > 3:
                size_info = self.translate(deliverySectionItems[2].text)
            else:
                size_info = ''

            material = deliverySectionItems[0].text

            care_instructions = self.translate(deliverySectionItems[1].text)

            description = self.translate(deliverySectionItems[-1].text).replace('\n', '').strip()

            name = self.translate(self.driver.find_element(By.XPATH, '//h1').text).strip()

            prices = list(map(float, self.driver.find_element(By.CSS_SELECTOR, '.productvariantcontent__price.js_pdpPrice').text.replace(' €', '').strip().split()))
            price = max(prices)
            price = self.get_cos_price(price)

            colors = [j.get_attribute('data-replaceurl') for j in self.driver.find_elements(By.XPATH, '//div[@class="swatchBox swatchBox--color "]/button')]
            article_num = self.driver.find_element(By.XPATH, '//span[@itemprop="productID"]').text
            try:
                regexp = re.compile(r'\d{1,3}% ([A-Za-z]+)')
                main_material = regexp.search(material).group(1)
            except:
                main_material = material

            main_material = self.translate(main_material)
            material = self.translate(material)
            if self.PARSE_TYPE == 'bags':
                for j in colors:
                    try:
                        self.driver.get(j)
                    except Exception:
                        continue
                    time.sleep(TIMEOUT)

                    photos = [i.get_attribute('data-splide-lazy') for i in
                              self.driver.find_elements(By.XPATH, '//img[@class="js_sliderThumbImg pdp__splideImg"]')]
                    main_photo = photos[0].replace('?width=60', '')
                    other_photo = []
                    for i in photos[1:]:
                        if i:
                            other_photo.append(i.replace('?width=60', ''))
                    other_photo = ','.join(other_photo)

                    c += 1

                    color = self.driver.find_element(By.CLASS_NAME, 'js-color').text

                    article = 'UNIQLO_' + article_num + '_' + color

                    rich = self.RICH.format(name, description, article_num, material, size_info)

                    self.COLUMNS['№'] = c
                    self.COLUMNS['Артикул*'] = article
                    self.COLUMNS['Название товара'] = name
                    try:
                        self.COLUMNS['Цена, руб.*'] = price
                    except:
                        self.COLUMNS['Цена, руб.*'] = 'Bad price'
                    self.COLUMNS['Ссылка на главное фото*'] = main_photo
                    self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                    self.COLUMNS['Название модели (для объединения в одну карточку)*'] = article_num
                    self.COLUMNS['Цвет товара'] = COLORS[color] if color in COLORS else 'разноцветный'
                    self.COLUMNS['Название цвета'] = self.translate(color)
                    self.COLUMNS['Страна-изготовитель'] = 'Турция'
                    self.COLUMNS['Материал'] = material
                    self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                    self.COLUMNS['Rich-контент JSON'] = rich

                    self.result.append(self.COLUMNS.copy())
            elif self.PARSE_TYPE == 'clothes':
                for j in colors:
                    try:
                        self.driver.get(j)
                    except Exception:
                        continue
                    time.sleep(TIMEOUT)

                    photos = [i.get_attribute('data-splide-lazy') for i in self.driver.find_elements(By.XPATH, '//img[@class="js_sliderThumbImg pdp__splideImg"]')]
                    main_photo = photos[0].replace('?width=60', '')
                    other_photo = []
                    for i in photos[1:]:
                        if i:
                            other_photo.append(i.replace('?width=60', ''))
                    other_photo = ','.join(other_photo)
                    sizes = self.driver.find_elements(By.CSS_SELECTOR, '.swatch.swatch--size')
                    for i in sizes:
                        c += 1

                        color = self.driver.find_element(By.CLASS_NAME, 'js-color').text

                        size = i.text
                        rich = self.RICH.format(name, description, article_num, material, size_info)
                        if 'jeans' not in self.CATEGORIE_URL:
                            article = 'UNIQLO_' + article_num + '_' + color + '_' + size

                            self.COLUMNS['№'] = c
                            self.COLUMNS['Артикул*'] = article
                            self.COLUMNS['Название товара'] = name
                            self.COLUMNS["Инструкция по уходу"] = care_instructions
                            try:
                                self.COLUMNS['Цена, руб.*'] = price
                            except:
                                self.COLUMNS['Цена, руб.*'] = 'Bad price'
                            self.COLUMNS['Ссылка на главное фото*'] = main_photo
                            self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                            self.COLUMNS['Объединить на одной карточке*'] = article_num
                            self.COLUMNS['Цвет товара*'] = COLORS[color] if color in COLORS else 'разноцветный'
                            if size.isdigit():
                                self.COLUMNS['Российский размер*'] = str(int(size) + 6)
                            else:
                                try:
                                    self.COLUMNS['Российский размер*'] = self.SIZES[size.upper()]
                                except:
                                    self.COLUMNS['Российский размер*'] = 'Bad size'  # Если размера нету в таблице размеров
                            self.COLUMNS['Размер производителя'] = size
                            self.COLUMNS['Название цвета'] = self.translate(color)
                            self.COLUMNS['Страна-изготовитель'] = 'Турция'
                            self.COLUMNS['Состав материала'] = material
                            self.COLUMNS['Материал'] = main_material
                            self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                            self.COLUMNS['Rich-контент JSON'] = rich
                        else:
                            lengths = self.driver.find_elements(By.CSS_SELECTOR, '.swatch.swatch--length')
                            for k in lengths:
                                length = k.text
                                article = 'UNIQLO_' + article_num + '_' + color + '_' + size + '_' + length
                                length = round(float(re.search('\d+', length)[0]) * 2.54)

                                self.COLUMNS['№'] = c
                                self.COLUMNS['Артикул*'] = article
                                self.COLUMNS['Название товара'] = name
                                self.COLUMNS["Инструкция по уходу"] = care_instructions
                                self.COLUMNS['Длина изделия, см'] = length
                                try:
                                    self.COLUMNS['Цена, руб.*'] = price
                                except:
                                    self.COLUMNS['Цена, руб.*'] = 'Bad price'
                                self.COLUMNS['Ссылка на главное фото*'] = main_photo
                                self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                                self.COLUMNS['Объединить на одной карточке*'] = article_num
                                self.COLUMNS['Цвет товара*'] = COLORS[color] if color in COLORS else 'разноцветный'
                                if size.isdigit():
                                    self.COLUMNS['Российский размер*'] = str(int(size) + 6)
                                else:
                                    try:
                                        self.COLUMNS['Российский размер*'] = self.SIZES[size.upper()]
                                    except:
                                        self.COLUMNS[
                                            'Российский размер*'] = 'Bad size'  # Если размера нету в таблице размеров
                                self.COLUMNS['Размер производителя'] = size
                                self.COLUMNS['Название цвета'] = self.translate(color)
                                self.COLUMNS['Страна-изготовитель'] = 'Турция'
                                self.COLUMNS['Состав материала'] = self.translate(material)
                                self.COLUMNS['Материал'] = self.translate(main_material)
                                self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                                self.COLUMNS['Rich-контент JSON'] = rich

                        self.result.append(self.COLUMNS.copy())

    def gPriceDict(self, key):
        return float(PRICE_TABLE[key])

    def get_cos_price(self, eur_price):
        cost_price = (float(eur_price) * self.gPriceDict("КОЭФ_КОНВЕРТАЦИИ") * self.gPriceDict(
            'КУРС_EUR_RUB')) + (self.DELIVERY_PRICE * self.gPriceDict('КУРС_БЕЛ.РУБ_РУБ') * self.gPriceDict(
            'КУРС_EUR_БЕЛ.РУБ'))
        final_price = (cost_price + self.gPriceDict('СРЕД_ЦЕН_ДОСТАВКИ')) / (
                1 - self.gPriceDict('НАЦЕНКА') - self.OZON_PRICE_MARKUP - self.gPriceDict(
            'ПРОЦЕНТЫ_НАЛОГ') - self.gPriceDict('ПРОЦЕНТЫ_ЭКВАЙРИНГ'))

        final_price = (final_price // 100 + 1) * 100 - 10
        return final_price

    def check_exists_by_xpath(self, xpath):
        try:
            self.driver.find_element(By.XPATH, xpath)
        except NoSuchElementException:
            return False
        return True

    def get_photo(self, url, name):
        r = requests.get(url, stream=True)
        if r.status_code == 200:
            with open(SAVE_PHOTO_PATH + name, 'wb') as f:
                r.raw.decode_content = True
                shutil.copyfileobj(r.raw, f)
            return 'http://' + HOST + '/COS_parser/' + SAVE_PHOTO_PATH + name
        else:
            return 'Bad photo'

    def translate(self, text):
        translator = Translator()
        while True:
            try:
                result = translator.translate(text, dest='ru')
                return result.text
            except:
                pass

    def save(self, result):
        wb = load_workbook(filename=f'{self.settings[CATEGORIE]["folder_path"]}/example.xlsx')
        ws = wb['Шаблон']
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        cols = []
        for col in alphabet:
            value = ws[col + '2'].value
            if value:
                cols.append(value)
        for col1 in alphabet:
            for col2 in alphabet:
                value = ws[col1 + col2 + '2'].value
                if value:
                    cols.append(value)

        for row in range(len(result)):
            for col in range(len(cols)):
                if cols[col] not in result[row]:
                    ws.cell(row=4 + row, column=1 + col).value = ''
                else:
                    ws.cell(row=4 + row, column=1 + col).value = result[row][cols[col]]

        wb.save(SAVE_XLSX_PATH + CATEGORIE + f"_{datetime.now()}.xlsx".replace(':', '.'))

    def sort_result(self):
        self.result.sort(key=lambda x: x['Артикул*'])
        for i in range(len(self.result)):
            self.result[i]['№'] = i + 1

    def load_settings(self):
        with open('settings.json', 'r', encoding='utf-8') as f:
            self.settings = json.load(f)
        self.CATEGORIE_URL = self.settings[CATEGORIE]['url']
        self.PARSE_TYPE = self.settings[CATEGORIE]['type_pars']
        self.DELIVERY_PRICE = float(self.settings[CATEGORIE]["ЦЕНА_ДОСТАВКИ_В_КАТЕГОРИИ"])
        self.OZON_PRICE_MARKUP = float(self.settings[CATEGORIE]["ПРОЦЕНТЫ_ОЗОН"])
        self.COLUMNS = self.load_module('columns').COLUMNS
        self.RICH = self.load_module('rich').RICH
        self.SIZES = self.load_module('sizes').SIZES
        self.TABLE_OF_SIZES = self.load_module('table_of_sizes').TABLE_OF_SIZES
        self.MATERIALS = MATERIALS
        self.COLORS = COLORS

    def load_module(self, name):
        spec = importlib.util.spec_from_file_location(name, self.settings[CATEGORIE]['folder_path'] + '/' + name + '.py')
        foo = importlib.util.module_from_spec(spec)
        sys.modules[name] = foo
        spec.loader.exec_module(foo)
        return foo

    def start(self):
        try:
            self.load_settings()
            print('--- START PARSING ---')
            self.parse()
            print('--- END PARSING ---')
        except Exception as e:
            error = self.driver.current_url + '\n' + traceback.format_exc() + '\n'
            print(error)
            with open('log.log', 'a') as f:
                f.write(error)
            with open('last.html', 'w') as f:
                f.write(self.driver.page_source)
        finally:
            self.sort_result()
            self.save(self.result)

            self.driver.close()
            self.driver.quit()


def main():
    parser = Parser()
    parser.start()


if __name__ == '__main__':
    if 'photo' not in os.listdir():
        os.mkdir('photo')
    if 'xlsx' not in os.listdir():
        os.mkdir('xlsx')
    if 'log.log' not in os.listdir():
        file = open('log.log', 'w')
        file.close()
    main()

